import openpyxl

#create a new workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Employees" # Rename sheet

#writing headers
sheet.append(["ID","Name","Age","Department"])

#writing data rows
data = [
    [1,"Daxa",25,"IT"],
    [2,"Ram",30,"HR"],
    [3,"Hanuman",27,"Finance"]
]

for row in data:
    sheet.append(row)

#save the file
# wb.save("FileHandling/Excel_Files/employees.xlsx")
# Save inside 'Excel_Files' folder
file_path = "Excel_Files/employees.xlsx"
wb.save(file_path)  
print("Excel file created successfully!")
print(f"File saved at: {file_path}")


 #reading 
wb = openpyxl.load_workbook(file_path)
sheet = wb.active
#print all rows
for row in sheet.iter_rows(values_only = True):
    print(row)


#append a new employee
sheet.append([4,"sam",29,"Marketing"])
#save changes
wb.save(file_path)
print("Data appended successfully!")
#print all rows
for row in sheet.iter_rows(values_only = True):
    print(row)

